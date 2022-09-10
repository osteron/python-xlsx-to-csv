import csv
from openpyxl import load_workbook


# поиск первой строки
def find_first_row(sheet_number):
    x = 1   # первая строка
    while '1' not in str(sheet_number.cell(row=x, column=9).value): # пока не найдем единицу в столбце ID
        x += 1
    return x


# поиск последней строки
def find_last_row(sheet_number, start):
    high = start    # первая строка со значением из find_first_row
    while sheet_number.cell(row=high, column=9).value is not None:  # пока не найдем None в столбце ID
        low = high
        high *= 2
    # двоичный поиск последней активной строки
    while (high-low) > 1:
        mid = (high+low) // 2
        if sheet_number.cell(row=mid, column=9).value is None:
            high = mid
        else:
            low = mid
    mid = (high + low) // 2
    return mid


# перебор всех листов и заполнение .csv файла
def parsing_sheets():
    print(f"Читаю документ CompanyReport...")
    wb = load_workbook(filename="./CompanyReport.xlsx") # путь к файлу
    print("Чтение завершено.")
    count_of_sheets = len(wb.sheetnames)    # количество листов
    csv_file = open('import.csv', 'w', newline='')  # путь к файлу csv
    with csv_file as file:
        writer = csv.writer(file)
        for i in range(1, count_of_sheets+1):   # перебор листов
            print(f'Обработка листа №{i}...')
            sheet = wb[f'Лист{i}']
            first_row = find_first_row(sheet)
            last_row = find_last_row(sheet, first_row)
            for k in range(first_row, last_row+1):  # проход всех активных строк
                # запись в .csv конкатенации строки из нужных столбцов
                writer.writerows(
                    [[str(sheet[f'I{k}'].value) + ";" + str(sheet[f'E{k}'].value) + ";" +
                      str(sheet[f'K{k}'].value) + ";" + str(sheet[f'H{k}'].value) + ";" +
                      str(sheet[f'J{k}'].value)]]
                )
            print(f'Обработка листа №{i} завершена.', end='\n')
    csv_file.close()
    wb.close()

if __name__ == "__main__":
    parsing_sheets()